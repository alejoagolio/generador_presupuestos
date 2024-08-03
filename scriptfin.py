import openpyxl
import streamlit as st

def obtener_precio(diametro_externo, diametro_interno, longitud, ratio_dolar):
    # Abre el archivo de Excel
    archivo_excel = openpyxl.load_workbook('Untitled spreadsheet.xlsx', data_only=True)

    # Selecciona la hoja de cálculo de precios
    hoja_precios = archivo_excel['List price']


    # Buscar el producto

    precio_final = None
    codigo = None

    medida = f'{diametro_externo}-{diametro_interno}'

    col = hoja_precios['H']

    for cel in col:
        if cel.value == medida:
            precio_1000 = hoja_precios[f'F{cel.row}'].value
            codigo = hoja_precios[f'G{cel.row}'].value
            break
    
    # Calcular el precio de oferta
    multiplicador = longitud/1000

    if multiplicador != 1:
        adicional_thordon = 0.1
    else:
        adicional_thordon = 0
    
    adicional_ml = 0.1

    precio_final = round(precio_1000 * ratio_dolar * (1+adicional_thordon) * (1+adicional_ml) * multiplicador, 2)

    return precio_final, codigo

def generar_presupuesto(productos, nombre_archivo):

    # Crear un nuevo libro de trabajo (workbook)
    libro = openpyxl.Workbook()

    # Seleccionar la hoja activa (por defecto, se crea una hoja llamada "Sheet")
    hoja_presupuesto = libro.active

    # Cambiar el nombre de la hoja activa
    hoja_presupuesto.title = 'Presupuesto'

    hoja_presupuesto['A1'] = 'Descripcion'
    hoja_presupuesto['B1'] = 'Cantidad'
    hoja_presupuesto['C1'] = 'Precio unitario'
    hoja_presupuesto['D1'] = 'Precio total'

    for producto in productos:
        
        codigo = producto[0]
        cantidad = producto[1]
        precio = producto[2]

        # Calcular el precio total

        total_prod = precio * cantidad

        fila = 1

        while hoja_presupuesto[f'A{fila}'].value != None:
            fila += 1

        # Agregar los datos del presupuesto
        hoja_presupuesto[f'A{fila}'] = codigo
        hoja_presupuesto[f'B{fila}'] = f"{cantidad}"
        hoja_presupuesto[f'C{fila}'] = f"{precio:.2f}"
        hoja_presupuesto[f'D{fila}'] = f"{total_prod:.2f}"


    # Calcular el total
    
    total = 0
    rw = 1

    col = hoja_presupuesto['D']
    for celda in col:
        rw += 1
        if celda.value != None and celda.row != 1:
            total += float(celda.value)
    
    hoja_presupuesto.cell(row = rw, column = 1, value = 'Total')
    hoja_presupuesto.cell(row = rw, column = 4, value = total)

    libro.save(f'{nombre_archivo}.xlsx')

    print(f"Presupuesto generado en '{nombre_archivo}.xlsx'.")

'''
def main():
    productos = []
    ratio_dolar = float(input("Ingrese la relacion USD/CAD: "))

    while True:
        diametro_externo = int(input("Ingrese el diámetro externo: "))
        diametro_interno = int(input("Ingrese el diámetro interno: "))
        longitud = int(input("Ingrese la longitud: "))
        cantidad = int(input("Ingrese la cantidad: "))

        precio, codigo = obtener_precio(diametro_externo, diametro_interno, longitud, ratio_dolar)
        if codigo is None or precio is None:
            print("No se pudieron obtener los valores calculados.")
            return

        productos.append((codigo, cantidad, precio))

        continuar = input("¿Desea agregar otro producto? (s/n): ")
        if continuar.lower() != 's':
            break

    generar_presupuesto(productos, "prueba1")
'''

def main():
    st.title("Generador de Presupuestos")

    ratio_dolar = st.number_input("Ingrese la relación USD/CAD:", min_value=0.0, step=0.01, format="%.2f")

    productos = []

    with st.form(key='my_form'):
        diametro_externo = st.number_input("Ingrese el diámetro externo:", min_value=0)
        diametro_interno = st.number_input("Ingrese el diámetro interno:", min_value=0)
        longitud = st.number_input("Ingrese la longitud:", min_value=0)
        cantidad = st.number_input("Ingrese la cantidad:", min_value=0)
        submit_button = st.form_submit_button(label='Agregar Producto')

        if submit_button:
            if diametro_externo <= 0 or diametro_interno <= 0 or longitud <= 0 or cantidad <= 0:
                st.error("Por favor, complete todos los campos correctamente.")
            else:
                precio, codigo = obtener_precio(diametro_externo, diametro_interno, longitud, ratio_dolar)
                if codigo is None or precio is None:
                    st.error("No se pudieron obtener los valores calculados.")
                else:
                    productos.append((codigo, cantidad, precio))
                    st.success(f"Producto agregado: {codigo} - Cantidad: {cantidad} - Precio: {precio:.2f}")

    if st.button("Generar Presupuesto"):
        if productos:
            generar_presupuesto(productos, "presupuesto")
            st.success("Presupuesto generado exitosamente.")
        else:
            st.error("No hay productos para generar el presupuesto.")

if __name__ == "__main__":
    main()
